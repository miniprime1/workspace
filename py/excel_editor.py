from openpyxl import *
from pandas import *
from tabulate import *

print("Excel Editor v1.0")
print("Copyright (c) 2020 miniprime1")
print("[Powered by OpenPyXL]\n")

print("Options")
print("1. Write Mode (w)")
print("2. Read Mode (r)")
print("3. Install Requirements (i)")
print("4. Exit (x)")
cc = input("Enter choice: ")

if cc == "1":
    path = input("\nEnter Save Path: ")
    wb = Workbook()
    ws = wb.active
    print("\n[Exit: enter 'EXIT' to exit]")
    print("[Cell: (Column)(Row)")
    while True:
        cell = input("\nEnter Cell: ")
        if cell=="EXIT": break
        value = input("Enter Value: ")
        if value=="EXIT": break
        ws[cell] = value
    wb.save(path)

elif cc == "2":
    path = input("\nEnter File Path: ")
    sheet = input("Enter Sheet Name: ")
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    df = DataFrame(ws.values)
    print("\n", end='')
    print(tabulate(ws.values, tablefmt="grid"))

elif cc == "3":
    exit()

# Install Requirements
#
# def install_package(package):
#     import os, sys
#     cmd = "".join((sys.executable, " ", "-m ", "pip ", "install ",  package))
#     os.system(cmd)
#
# install_package("openpyxl")
# install_package("pandas")
# install_package("tabulate")